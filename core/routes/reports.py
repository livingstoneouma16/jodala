from flask import Blueprint, request, jsonify, render_template, send_file, Response
from datetime import date, timedelta
import csv
import io

from core.database import get_db
from core.auth import login_required, get_current_user
from core.serializers import loan_public, repayment_public, member_public, member_full_name, client_full_name, client_public

reports_bp = Blueprint('reports', __name__)


def _write_chart_data_sheet(wb, sheet_name, categories, series):
    """Writes a small data table on its own sheet (kept out of the way after
    the main report sheet) and returns (data_ws, min_col, max_col, min_row,
    max_row) refs openpyxl charts need. `series` is a list of (label, values)
    tuples, all the same length as `categories`."""
    data_ws = wb.create_sheet(sheet_name)
    data_ws.cell(row=1, column=1, value='Category')
    for col, (label, _values) in enumerate(series, 2):
        data_ws.cell(row=1, column=col, value=label)
    for row, cat in enumerate(categories, 2):
        data_ws.cell(row=row, column=1, value=cat)
    for col, (_label, values) in enumerate(series, 2):
        for row, value in enumerate(values, 2):
            data_ws.cell(row=row, column=col, value=value)
    data_ws.sheet_state = 'hidden'
    return data_ws, 1, len(series) + 1, 1, len(categories) + 1


def _add_pie_chart(wb, ws, anchor, title, categories, values, sheet_name):
    from openpyxl.chart import PieChart, Reference

    data_ws, _, _, _, r2 = _write_chart_data_sheet(wb, sheet_name, categories, [(title, values)])
    chart = PieChart()
    chart.title = title
    chart.height, chart.width = 8, 12
    data_ref = Reference(data_ws, min_col=2, min_row=1, max_row=r2)
    cats_ref = Reference(data_ws, min_col=1, min_row=2, max_row=r2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor)


def _add_bar_chart(wb, ws, anchor, title, categories, series, sheet_name, y_title=''):
    """series: list of (label, values) tuples for grouped bars."""
    from openpyxl.chart import BarChart, Reference

    data_ws, _, c2, _, r2 = _write_chart_data_sheet(wb, sheet_name, categories, series)
    chart = BarChart()
    chart.type = 'col'
    chart.title = title
    chart.y_axis.title = y_title
    chart.height, chart.width = 8, 14
    data_ref = Reference(data_ws, min_col=2, max_col=c2, min_row=1, max_row=r2)
    cats_ref = Reference(data_ws, min_col=1, min_row=2, max_row=r2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor)


def _add_line_chart(wb, ws, anchor, title, categories, values, sheet_name, y_title=''):
    from openpyxl.chart import LineChart, Reference

    data_ws, _, _, _, r2 = _write_chart_data_sheet(wb, sheet_name, categories, [(title, values)])
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.height, chart.width = 8, 14
    data_ref = Reference(data_ws, min_col=2, min_row=1, max_row=r2)
    cats_ref = Reference(data_ws, min_col=1, min_row=2, max_row=r2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor)


def _loan_join_sql(where_sql=''):
    return f"""SELECT loans.*,
                      COALESCE(
                          TRIM(members.first_name || ' ' || COALESCE(members.middle_name, '') || ' ' || members.last_name),
                          TRIM(clients.first_name || ' ' || clients.last_name)
                      ) AS borrower_name,
                      COALESCE(members.phone, clients.phone) AS borrower_phone,
                      COALESCE(members.region, clients.region) AS region,
                      loan_products.name AS product_name,
                      officers.full_name AS loan_officer_name
               FROM loans
               LEFT JOIN members ON members.id = loans.member_id
               LEFT JOIN clients ON clients.id = loans.client_id
               LEFT JOIN loan_products ON loan_products.id = loans.product_id
               LEFT JOIN users officers ON officers.id = loans.loan_officer_id{where_sql}"""


@reports_bp.route('/')
@login_required
def index():
    return render_template('reports/index.html', user=get_current_user())


@reports_bp.route('/api/loan-report')
@login_required
def loan_report():
    status = request.args.get('status', '')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    where, params = [], []
    if status:
        where.append("loans.status = %s"); params.append(status)
    if date_from:
        where.append("loans.application_date >= %s"); params.append(date_from)
    if date_to:
        where.append("loans.application_date <= %s"); params.append(date_to)

    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    loans = get_db().execute(_loan_join_sql(where_sql), tuple(params)).fetchall()

    by_status = {}
    for l in loans:
        by_status[l['status']] = by_status.get(l['status'], 0) + 1

    summary = {
        'total': len(loans),
        'total_principal': sum(l['principal_amount'] for l in loans),
        'total_disbursed': sum(l['amount_disbursed'] for l in loans),
        'total_outstanding': sum(l['outstanding_balance'] for l in loans),
        'total_collected': sum(l['total_paid'] for l in loans),
    }

    return jsonify({
        'loans': [loan_public(l) for l in loans[:100]],
        'summary': {k: round(v, 2) for k, v in summary.items()},
        'by_status': by_status
    })


@reports_bp.route('/api/portfolio-source-report')
@login_required
def portfolio_source_report():
    """Imported (via Portfolio Import) vs originated-in-app breakdown --
    the is_imported flag exists on every loan (migration 27) but nothing
    surfaced it anywhere until now. Useful right after a migration to spot-
    check the imported book, and afterwards to see how much of the active
    portfolio still traces back to the old system vs new business written
    directly in this app."""
    loans = get_db().execute(_loan_join_sql()).fetchall()

    def _bucket(rows):
        return {
            'count': len(rows),
            'total_principal': round(sum(l['principal_amount'] for l in rows), 2),
            'total_outstanding': round(sum(l['outstanding_balance'] for l in rows), 2),
            'total_collected': round(sum(l['total_paid'] for l in rows), 2),
            'by_status': {s: sum(1 for l in rows if l['status'] == s) for s in
                          {l['status'] for l in rows}},
        }

    imported = [l for l in loans if l['is_imported']]
    originated = [l for l in loans if not l['is_imported']]

    batches = get_db().execute(
        """SELECT b.*, u.full_name AS created_by_name
           FROM loan_import_batches b LEFT JOIN users u ON u.id = b.created_by
           ORDER BY b.created_at DESC"""
    ).fetchall()

    return jsonify({
        'imported': _bucket(imported),
        'originated': _bucket(originated),
        'batches': [dict(b) for b in batches],
    })


@reports_bp.route('/api/arrears-report')
@login_required
def arrears_report():
    today = date.today()
    overdue = get_db().execute(
        """SELECT loan_schedules.*,
                  loans.loan_number, loans.status AS loan_status, loans.outstanding_balance,
                  loans.member_id, loans.client_id
           FROM loan_schedules
           LEFT JOIN loans ON loans.id = loan_schedules.loan_id
           WHERE loan_schedules.due_date < %s AND loan_schedules.status IN ('pending', 'partial')""",
        (today.isoformat(),)
    ).fetchall()

    active_overdue = [s for s in overdue if s['loan_status'] == 'active']

    member_ids = {s['member_id'] for s in active_overdue if s['member_id']}
    client_ids = {s['client_id'] for s in active_overdue if s['client_id']}

    members_by_id = {}
    if member_ids:
        placeholders = ','.join(['%s'] * len(member_ids))
        for m in get_db().execute(f"SELECT * FROM members WHERE id IN ({placeholders})", tuple(member_ids)).fetchall():
            members_by_id[m['id']] = member_full_name(m)

    clients_by_id = {}
    if client_ids:
        placeholders = ','.join(['%s'] * len(client_ids))
        for c in get_db().execute(f"SELECT * FROM clients WHERE id IN ({placeholders})", tuple(client_ids)).fetchall():
            clients_by_id[c['id']] = client_full_name(c)

    result = []
    for s in active_overdue:
        days_overdue = (today - date.fromisoformat(s['due_date'])).days
        borrower = members_by_id.get(s['member_id']) or clients_by_id.get(s['client_id']) or 'N/A'

        result.append({
            'loan_number': s['loan_number'],
            'borrower': borrower,
            'member_id': s['member_id'],
            'client_id': s['client_id'],
            'installment': s['installment_number'],
            'due_date': s['due_date'],
            'amount_due': s['total_due'] - s['total_paid'],
            'days_overdue': days_overdue,
            'outstanding_balance': s['outstanding_balance']
        })

    result.sort(key=lambda x: x['days_overdue'], reverse=True)
    total_arrears = sum(r['amount_due'] for r in result)
    return jsonify({'arrears': result, 'total_arrears': round(total_arrears, 2), 'count': len(result)})


@reports_bp.route('/api/collection-report')
@login_required
def collection_report():
    date_from = request.args.get('date_from', date.today().replace(day=1).isoformat())
    date_to = request.args.get('date_to', date.today().isoformat())

    repayments = get_db().execute(
        """SELECT repayments.*, loans.loan_number FROM repayments
           LEFT JOIN loans ON loans.id = repayments.loan_id
           WHERE repayments.payment_date >= %s AND repayments.payment_date <= %s
           ORDER BY repayments.payment_date DESC""",
        (date_from, date_to)
    ).fetchall()

    total = sum(r['amount'] for r in repayments)
    by_method = {}
    for r in repayments:
        by_method[r['payment_method']] = by_method.get(r['payment_method'], 0) + r['amount']

    return jsonify({
        'repayments': [repayment_public(r) for r in repayments],
        'total': round(total, 2),
        'by_method': {k: round(v, 2) for k, v in by_method.items()},
        'count': len(repayments)
    })


@reports_bp.route('/api/member-report')
@login_required
def member_report():
    """Combined borrower report: members and clients are two separate
    tables (see core/database.py) but both represent people loans are
    issued to, so this report merges them into one list the same way
    loans already resolve a borrower from either table (see
    _loan_join_sql above). Each row carries a borrower_type field so the
    UI/exports can still tell them apart."""
    members = get_db().execute("SELECT * FROM members ORDER BY created_at DESC").fetchall()
    clients = get_db().execute("SELECT * FROM clients ORDER BY created_at DESC").fetchall()

    by_status, by_region = {}, {}
    borrowers = []

    for m in members:
        by_status[m['status']] = by_status.get(m['status'], 0) + 1
        if m['region']:
            by_region[m['region']] = by_region.get(m['region'], 0) + 1
        row = member_public(m)
        row['borrower_type'] = 'member'
        row['number'] = row['member_number']
        borrowers.append(row)

    for c in clients:
        by_status[c['status']] = by_status.get(c['status'], 0) + 1
        if c['region']:
            by_region[c['region']] = by_region.get(c['region'], 0) + 1
        row = client_public(c)
        row['borrower_type'] = 'client'
        row['number'] = row['client_number']
        borrowers.append(row)

    borrowers.sort(key=lambda b: b['created_at'] or '', reverse=True)

    return jsonify({
        'members': borrowers[:200],
        'total': len(borrowers),
        'member_count': len(members),
        'client_count': len(clients),
        'by_status': by_status,
        'by_region': by_region
    })


def _compute_regional_performance(month=None):
    """Portfolio performance broken down by borrower region: reach (members),
    volume (loan counts/principal/disbursed), and health (outstanding,
    collected, arrears, PAR%, collection rate%). A loan's region comes from
    its member/client (see _loan_join_sql) -- loans themselves don't carry a
    region directly. Plain function (no @login_required) so both the JSON
    endpoint and the Excel/CSV exports can call it directly without an extra
    auth check or a jsonify->get_json round trip.

    `month` scopes the report to a single calendar month ('YYYY-MM'),
    defaulting to the current month -- this is a monthly report, not a
    lifetime snapshot. It's applied per-metric rather than as one blanket
    date filter, since "this month" means something different for each:
      - member_count: members whose account was created in that month
        (region reach gained that month, not the region's all-time headcount)
      - loan_count/active_loan_count/total_principal/total_disbursed:
        loans disbursed in that month
      - total_collected: repayments actually received in that month
      - total_outstanding/par_outstanding/par_pct: as of month-end (a
        balance is a point-in-time figure, not something that accrues
        "during" a month, so this uses the snapshot at the end of the
        selected month rather than summing disbursed-that-month balances)
      - arrears_amount: installments overdue as of month-end
    """
    today = date.today()
    if month:
        year, mon = (int(x) for x in month.split('-'))
    else:
        year, mon = today.year, today.month

    month_start = date(year, mon, 1)
    month_end = date(year + 1, 1, 1) if mon == 12 else date(year, mon + 1, 1)
    # For "as of month-end" balance/PAR figures: if the selected month is
    # the current month, month-end hasn't happened yet, so "as of" has to
    # mean "as of today" instead of a future date.
    as_of = min(month_end, today + timedelta(days=1))

    loans = get_db().execute(
        _loan_join_sql() + " WHERE loans.disbursement_date >= %s AND loans.disbursement_date < %s "
                            "ORDER BY loans.created_at DESC",
        (month_start.isoformat(), month_end.isoformat())
    ).fetchall()

    # All loans (regardless of disbursement month) are needed for the
    # as-of-month-end balance/PAR figures -- a loan disbursed last month
    # still contributes to this month's outstanding balance.
    all_loans = get_db().execute(_loan_join_sql()).fetchall()

    # Overdue installments as of month-end, joined just enough to resolve a
    # region without an N+1 query per row (member_regions/client_regions are
    # prefetched once below instead of queried per overdue row).
    overdue = get_db().execute(
        """SELECT loan_schedules.loan_id, loan_schedules.total_due, loan_schedules.total_paid,
                  loans.status AS loan_status, loans.member_id, loans.client_id
           FROM loan_schedules
           LEFT JOIN loans ON loans.id = loan_schedules.loan_id
           WHERE loan_schedules.due_date < %s AND loan_schedules.status IN ('pending', 'partial')""",
        (as_of.isoformat(),)
    ).fetchall()

    # Repayments received within the month, for the total_collected figure.
    collections = get_db().execute(
        """SELECT loans.member_id, loans.client_id, repayments.amount_paid
           FROM repayments
           LEFT JOIN loans ON loans.id = repayments.loan_id
           WHERE repayments.payment_date >= %s AND repayments.payment_date < %s""",
        (month_start.isoformat(), month_end.isoformat())
    ).fetchall()

    member_regions = {m['id']: m['region'] for m in get_db().execute("SELECT id, region FROM members").fetchall()}
    client_regions = {c['id']: c['region'] for c in get_db().execute("SELECT id, region FROM clients").fetchall()}
    members_this_month = get_db().execute(
        "SELECT id, region FROM members WHERE created_at >= %s AND created_at < %s",
        (month_start.isoformat(), month_end.isoformat())
    ).fetchall()
    clients_this_month = get_db().execute(
        "SELECT id, region FROM clients WHERE created_at >= %s AND created_at < %s",
        (month_start.isoformat(), month_end.isoformat())
    ).fetchall()

    regions = {}

    def bucket(name):
        key = name or 'Unassigned'
        return regions.setdefault(key, {
            'region': key, 'member_count': 0, 'loan_count': 0, 'active_loan_count': 0,
            'total_principal': 0.0, 'total_disbursed': 0.0, 'total_outstanding': 0.0,
            'total_collected': 0.0, 'par_outstanding': 0.0, 'arrears_amount': 0.0,
        })

    # Loans with at least one overdue installment -- used below to attribute
    # outstanding balance to "at risk" (PAR) per region.
    overdue_loan_ids = set()
    for s in overdue:
        if s['loan_status'] != 'active':
            continue
        overdue_loan_ids.add(s['loan_id'])
        region = member_regions.get(s['member_id']) if s['member_id'] else client_regions.get(s['client_id'])
        bucket(region)['arrears_amount'] += (s['total_due'] - s['total_paid'])

    # Volume figures (loan_count, principal, disbursed) are scoped to loans
    # disbursed in the selected month.
    for l in loans:
        b = bucket(l['region'])
        b['loan_count'] += 1
        b['total_principal'] += l['principal_amount'] or 0
        b['total_disbursed'] += l['amount_disbursed'] or 0
        if l['status'] == 'active':
            b['active_loan_count'] += 1

    # Balance/PAR figures are as-of-month-end, so they're computed from the
    # full loan book rather than just this month's disbursements.
    for l in all_loans:
        b = bucket(l['region'])
        b['total_outstanding'] += l['outstanding_balance'] or 0
        if l['status'] == 'active' and l['id'] in overdue_loan_ids:
            b['par_outstanding'] += l['outstanding_balance'] or 0

    for c in collections:
        region = member_regions.get(c['member_id']) if c['member_id'] else client_regions.get(c['client_id'])
        bucket(region)['total_collected'] += c['amount_paid'] or 0

    for m in members_this_month:
        bucket(m['region'])['member_count'] += 1
    for c in clients_this_month:
        bucket(c['region'])['member_count'] += 1

    result = []
    for data in regions.values():
        outstanding = data['total_outstanding']
        # Collection rate uses collected / (collected + outstanding) rather
        # than collected / total_repayable, since total_repayable includes
        # interest on loans not yet fully scheduled/disbursed and would
        # understate the rate for a young, fast-growing region.
        collectible_base = data['total_collected'] + outstanding
        data['par_pct'] = round(data['par_outstanding'] / outstanding * 100, 2) if outstanding else 0.0
        data['collection_rate_pct'] = round(data['total_collected'] / collectible_base * 100, 2) if collectible_base else 0.0
        for key in ('total_principal', 'total_disbursed', 'total_outstanding', 'total_collected', 'arrears_amount'):
            data[key] = round(data[key], 2)
        del data['par_outstanding']
        result.append(data)

    result.sort(key=lambda r: r['total_outstanding'], reverse=True)

    totals = {
        'region_count': len(result),
        'member_count': sum(r['member_count'] for r in result),
        'loan_count': sum(r['loan_count'] for r in result),
        'total_outstanding': round(sum(r['total_outstanding'] for r in result), 2),
        'total_collected': round(sum(r['total_collected'] for r in result), 2),
        'arrears_amount': round(sum(r['arrears_amount'] for r in result), 2),
    }

    return {'month': f'{year:04d}-{mon:02d}', 'regions': result, 'totals': totals}


@reports_bp.route('/api/regional-performance')
@login_required
def regional_performance():
    return jsonify(_compute_regional_performance(request.args.get('month')))


@reports_bp.route('/api/export/loans/excel')
@login_required
def export_loans_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    loans = get_db().execute(_loan_join_sql() + " ORDER BY loans.created_at DESC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Loans Report"

    headers = ['Loan No', 'Borrower', 'Phone', 'Product', 'Principal', 'Interest Rate',
               'Total Repayable', 'Outstanding', 'Total Paid', 'Status',
               'Application Date', 'Disbursement Date', 'Due Date']

    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 15

    for row, loan in enumerate(loans, 2):
        ws.cell(row=row, column=1, value=loan['loan_number'])
        ws.cell(row=row, column=2, value=loan['borrower_name'] or 'N/A')
        ws.cell(row=row, column=3, value=loan['borrower_phone'] or '')
        ws.cell(row=row, column=4, value=loan['product_name'] or '')
        ws.cell(row=row, column=5, value=loan['principal_amount'])
        ws.cell(row=row, column=6, value=f"{loan['interest_rate']}%")
        ws.cell(row=row, column=7, value=loan['total_repayable'])
        ws.cell(row=row, column=8, value=loan['outstanding_balance'])
        ws.cell(row=row, column=9, value=loan['total_paid'])
        ws.cell(row=row, column=10, value=(loan['status'] or '').upper())
        ws.cell(row=row, column=11, value=loan['application_date'] or '')
        ws.cell(row=row, column=12, value=loan['disbursement_date'] or '')
        ws.cell(row=row, column=13, value=loan['expected_end_date'] or '')

    by_status = {}
    for loan in loans:
        by_status[loan['status']] = by_status.get(loan['status'], 0) + 1

    if by_status:
        _add_pie_chart(wb, ws, 'O2', 'Loans by Status',
                        list(by_status.keys()), list(by_status.values()), '_chart_status')

    if loans:
        totals = ['Principal', 'Outstanding', 'Total Paid']
        total_values = [
            sum(l['principal_amount'] for l in loans),
            sum(l['outstanding_balance'] for l in loans),
            sum(l['total_paid'] for l in loans),
        ]
        _add_bar_chart(wb, ws, 'O20', 'Principal, Outstanding & Collected',
                        totals, [('Amount (Ksh)', total_values)], '_chart_amounts', y_title='Ksh')

    by_product = {}
    by_officer = {}
    by_month = {}
    for loan in loans:
        product = loan['product_name'] or 'Unassigned'
        by_product[product] = by_product.get(product, 0) + 1

        officer = loan['loan_officer_name'] or 'Unassigned'
        by_officer[officer] = by_officer.get(officer, 0) + loan['principal_amount']

        disb = loan['disbursement_date']
        if disb:
            month = str(disb)[:7]  # YYYY-MM
            by_month[month] = by_month.get(month, 0) + loan['principal_amount']

    if by_product:
        _add_bar_chart(wb, ws, 'O38', 'Loans by Product', list(by_product.keys()),
                        [('Loan Count', list(by_product.values()))], '_chart_product')

    if by_officer:
        _add_bar_chart(wb, ws, 'O56', 'Principal Disbursed by Loan Officer', list(by_officer.keys()),
                        [('Principal (Ksh)', list(by_officer.values()))], '_chart_officer', y_title='Ksh')

    if by_month:
        months_sorted = sorted(by_month.keys())
        _add_line_chart(wb, ws, 'O74', 'Disbursement Trend', months_sorted,
                         [by_month[m] for m in months_sorted], '_chart_disbursement', y_title='Ksh')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='loans_report.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/api/export/members/excel')
@login_required
def export_members_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    members = get_db().execute("SELECT * FROM members ORDER BY created_at DESC").fetchall()
    clients = get_db().execute("SELECT * FROM clients ORDER BY created_at DESC").fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Members & Clients"

    headers = ['Type', 'No.', 'First Name', 'Last Name', 'Phone', 'Email',
               'Region', 'Occupation', 'Status', 'Joined']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    row = 2
    by_status, by_region, by_month = {}, {}, {}
    for m in members:
        ws.cell(row=row, column=1, value='Member')
        ws.cell(row=row, column=2, value=m['member_number'])
        ws.cell(row=row, column=3, value=m['first_name'])
        ws.cell(row=row, column=4, value=m['last_name'])
        ws.cell(row=row, column=5, value=m['phone'])
        ws.cell(row=row, column=6, value=m['email'])
        ws.cell(row=row, column=7, value=m['region'])
        ws.cell(row=row, column=8, value=m['occupation'])
        ws.cell(row=row, column=9, value=m['status'])
        ws.cell(row=row, column=10, value=m['created_at'])
        by_status[m['status']] = by_status.get(m['status'], 0) + 1
        if m['region']:
            by_region[m['region']] = by_region.get(m['region'], 0) + 1
        if m['created_at']:
            month = str(m['created_at'])[:7]
            by_month[month] = by_month.get(month, 0) + 1
        row += 1

    for c in clients:
        ws.cell(row=row, column=1, value='Client')
        ws.cell(row=row, column=2, value=c['client_number'])
        ws.cell(row=row, column=3, value=c['first_name'])
        ws.cell(row=row, column=4, value=c['last_name'])
        ws.cell(row=row, column=5, value=c['phone'])
        ws.cell(row=row, column=6, value=c['email'])
        ws.cell(row=row, column=7, value=c['region'])
        ws.cell(row=row, column=8, value=c['occupation'])
        ws.cell(row=row, column=9, value=c['status'])
        ws.cell(row=row, column=10, value=c['created_at'])
        by_status[c['status']] = by_status.get(c['status'], 0) + 1
        if c['region']:
            by_region[c['region']] = by_region.get(c['region'], 0) + 1
        if c['created_at']:
            month = str(c['created_at'])[:7]
            by_month[month] = by_month.get(month, 0) + 1
        row += 1

    if by_status:
        _add_pie_chart(wb, ws, 'M2', 'Members & Clients by Status',
                        list(by_status.keys()), list(by_status.values()), '_chart_status')
    if by_region:
        _add_bar_chart(wb, ws, 'M20', 'Members & Clients by Region',
                        list(by_region.keys()), [('Count', list(by_region.values()))], '_chart_region')
    if by_month:
        months_sorted = sorted(by_month.keys())
        _add_line_chart(wb, ws, 'M38', 'Growth Over Time (New Joins per Month)', months_sorted,
                         [by_month[m] for m in months_sorted], '_chart_growth', y_title='New Joins')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='members_report.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


_REGIONAL_HEADERS = ['Region', 'Members', 'Loans', 'Active Loans', 'Total Principal',
                     'Total Disbursed', 'Outstanding', 'Total Collected', 'Arrears',
                     'PAR %', 'Collection Rate %']


def _regional_rows(regions):
    return [
        [
            r['region'], r['member_count'], r['loan_count'], r['active_loan_count'],
            r['total_principal'], r['total_disbursed'], r['total_outstanding'],
            r['total_collected'], r['arrears_amount'], r['par_pct'], r['collection_rate_pct'],
        ]
        for r in regions
    ]


@reports_bp.route('/api/export/regional/excel')
@login_required
def export_regional_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    data = _compute_regional_performance(request.args.get('month'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Regional Performance"

    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.cell(row=1, column=1, value=f"Regional Performance -- {data['month']}").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_REGIONAL_HEADERS))

    for col, header in enumerate(_REGIONAL_HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 16

    for row, values in enumerate(_regional_rows(data['regions']), 3):
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value)

    regions = data['regions']
    if regions:
        names = [r['region'] for r in regions]
        # Default chart height is ~8 rows tall; stack each chart below the
        # last with a margin instead of hardcoded rows, so charts can't
        # overlap as the region count grows.
        row_1 = 3
        row_2 = row_1 + 18
        row_3 = row_2 + 18
        row_4 = row_3 + 18
        _add_bar_chart(wb, ws, f'M{row_1}', 'Outstanding vs Collected by Region', names,
                        [('Outstanding', [r['total_outstanding'] for r in regions]),
                         ('Collected', [r['total_collected'] for r in regions])],
                        '_chart_outstanding', y_title='Ksh')
        _add_bar_chart(wb, ws, f'M{row_2}', 'Portfolio at Risk (PAR %) by Region', names,
                        [('PAR %', [r['par_pct'] for r in regions])],
                        '_chart_par', y_title='%')
        _add_bar_chart(wb, ws, f'M{row_3}', 'Loan Count by Region', names,
                        [('Loans', [r['loan_count'] for r in regions])],
                        '_chart_loan_count')
        avg_loan_size = [
            (r['total_principal'] / r['loan_count']) if r['loan_count'] else 0
            for r in regions
        ]
        _add_bar_chart(wb, ws, f'M{row_4}', 'Average Loan Size by Region', names,
                        [('Average Principal (Ksh)', avg_loan_size)], '_chart_avg_size', y_title='Ksh')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                      download_name=f"regional_performance_{data['month']}.xlsx",
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _csv_response(headers, rows, download_name):
    """Build a CSV download from a header row + list of value tuples/lists.
    Uses Python's stdlib csv module (no new dependency) with \\r\\n line
    endings and full quoting, matching what Excel/Sheets expect on import --
    a plain '\\n'.join(','.join(...)) would break on any field containing a
    comma, quote, or newline (e.g. a member's notes field)."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    writer.writerows(rows)
    csv_bytes = output.getvalue().encode('utf-8-sig')  # BOM so Excel on Windows detects UTF-8 correctly
    return Response(
        csv_bytes,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={download_name}'}
    )


@reports_bp.route('/api/export/arrears/excel')
@login_required
def export_arrears_excel():
    """Excel sibling of /api/arrears-report -- rebuilds the same overdue-
    installment query rather than calling the view function directly, so
    this stays a plain data export independent of the JSON response shape."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    today = date.today()
    overdue = get_db().execute(
        """SELECT loan_schedules.*,
                  loans.loan_number, loans.status AS loan_status, loans.outstanding_balance,
                  loans.member_id, loans.client_id
           FROM loan_schedules
           LEFT JOIN loans ON loans.id = loan_schedules.loan_id
           WHERE loan_schedules.due_date < %s AND loan_schedules.status IN ('pending', 'partial')""",
        (today.isoformat(),)
    ).fetchall()
    active_overdue = [s for s in overdue if s['loan_status'] == 'active']

    member_ids = {s['member_id'] for s in active_overdue if s['member_id']}
    client_ids = {s['client_id'] for s in active_overdue if s['client_id']}
    members_by_id, clients_by_id = {}, {}
    if member_ids:
        placeholders = ','.join(['%s'] * len(member_ids))
        for m in get_db().execute(f"SELECT * FROM members WHERE id IN ({placeholders})", tuple(member_ids)).fetchall():
            members_by_id[m['id']] = member_full_name(m)
    if client_ids:
        placeholders = ','.join(['%s'] * len(client_ids))
        for c in get_db().execute(f"SELECT * FROM clients WHERE id IN ({placeholders})", tuple(client_ids)).fetchall():
            clients_by_id[c['id']] = client_full_name(c)

    rows = []
    for s in active_overdue:
        days_overdue = (today - date.fromisoformat(s['due_date'])).days
        borrower = members_by_id.get(s['member_id']) or clients_by_id.get(s['client_id']) or 'N/A'
        rows.append({
            'loan_number': s['loan_number'], 'borrower': borrower,
            'installment': s['installment_number'], 'due_date': s['due_date'],
            'amount_due': s['total_due'] - s['total_paid'], 'days_overdue': days_overdue,
            'outstanding_balance': s['outstanding_balance']
        })
    rows.sort(key=lambda x: x['days_overdue'], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Arrears Report"

    headers = ['Loan No', 'Borrower', 'Installment #', 'Due Date', 'Amount Due',
               'Days Overdue', 'Outstanding Balance']
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 16

    aging_buckets = {'1-30 days': 0, '31-60 days': 0, '61-90 days': 0, '90+ days': 0}
    for row_num, r in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=r['loan_number'])
        ws.cell(row=row_num, column=2, value=r['borrower'])
        ws.cell(row=row_num, column=3, value=r['installment'])
        ws.cell(row=row_num, column=4, value=r['due_date'])
        ws.cell(row=row_num, column=5, value=r['amount_due'])
        ws.cell(row=row_num, column=6, value=r['days_overdue'])
        ws.cell(row=row_num, column=7, value=r['outstanding_balance'])

        d = r['days_overdue']
        if d <= 30:
            aging_buckets['1-30 days'] += r['amount_due']
        elif d <= 60:
            aging_buckets['31-60 days'] += r['amount_due']
        elif d <= 90:
            aging_buckets['61-90 days'] += r['amount_due']
        else:
            aging_buckets['90+ days'] += r['amount_due']

    if rows:
        _add_bar_chart(wb, ws, 'I2', 'Arrears Aging (Amount Due)', list(aging_buckets.keys()),
                        [('Amount Due (Ksh)', list(aging_buckets.values()))], '_chart_aging', y_title='Ksh')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='arrears_report.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/api/export/loans/csv')
@login_required
def export_loans_csv():
    loans = get_db().execute(_loan_join_sql() + " ORDER BY loans.created_at DESC").fetchall()

    headers = ['Loan No', 'Borrower', 'Phone', 'Product', 'Principal', 'Interest Rate',
               'Total Repayable', 'Outstanding', 'Total Paid', 'Status',
               'Application Date', 'Disbursement Date', 'Due Date']

    rows = [
        [
            loan['loan_number'],
            loan['borrower_name'] or 'N/A',
            loan['borrower_phone'] or '',
            loan['product_name'] or '',
            loan['principal_amount'],
            f"{loan['interest_rate']}%",
            loan['total_repayable'],
            loan['outstanding_balance'],
            loan['total_paid'],
            (loan['status'] or '').upper(),
            loan['application_date'] or '',
            loan['disbursement_date'] or '',
            loan['expected_end_date'] or '',
        ]
        for loan in loans
    ]

    return _csv_response(headers, rows, 'loans_report.csv')


@reports_bp.route('/api/export/members/csv')
@login_required
def export_members_csv():
    members = get_db().execute("SELECT * FROM members ORDER BY created_at DESC").fetchall()
    clients = get_db().execute("SELECT * FROM clients ORDER BY created_at DESC").fetchall()

    headers = ['Type', 'No.', 'First Name', 'Last Name', 'Phone', 'Email',
               'Region', 'Occupation', 'Status', 'Joined']

    rows = [
        [
            'Member', m['member_number'], m['first_name'], m['last_name'], m['phone'],
            m['email'], m['region'], m['occupation'], m['status'], m['created_at'],
        ]
        for m in members
    ] + [
        [
            'Client', c['client_number'], c['first_name'], c['last_name'], c['phone'],
            c['email'], c['region'], c['occupation'], c['status'], c['created_at'],
        ]
        for c in clients
    ]

    return _csv_response(headers, rows, 'members_report.csv')


@reports_bp.route('/api/export/collections/excel')
@login_required
def export_collections_excel():
    """Excel sibling of /api/collection-report -- same date_from/date_to
    filters and query, so the download always matches what's on screen."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    date_from = request.args.get('date_from', date.today().replace(day=1).isoformat())
    date_to = request.args.get('date_to', date.today().isoformat())

    repayments = get_db().execute(
        """SELECT repayments.*, loans.loan_number, officers.full_name AS collector_name FROM repayments
           LEFT JOIN loans ON loans.id = repayments.loan_id
           LEFT JOIN users officers ON officers.id = repayments.collected_by
           WHERE repayments.payment_date >= %s AND repayments.payment_date <= %s
           ORDER BY repayments.payment_date DESC""",
        (date_from, date_to)
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Collections Report"

    headers = ['Loan No', 'Amount', 'Payment Method', 'Payment Date', 'Reference']
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 16

    by_method, by_day, by_officer, by_day_method = {}, {}, {}, {}
    for row, r in enumerate(repayments, 2):
        ws.cell(row=row, column=1, value=r['loan_number'])
        ws.cell(row=row, column=2, value=r['amount'])
        ws.cell(row=row, column=3, value=(r['payment_method'] or '').replace('_', ' ').title())
        ws.cell(row=row, column=4, value=r['payment_date'])
        ws.cell(row=row, column=5, value=r['reference_number'] or '')
        by_method[r['payment_method']] = by_method.get(r['payment_method'], 0) + r['amount']
        by_day[r['payment_date']] = by_day.get(r['payment_date'], 0) + r['amount']
        officer = r['collector_name'] or 'Unassigned'
        by_officer[officer] = by_officer.get(officer, 0) + r['amount']

    if by_method:
        _add_pie_chart(wb, ws, 'H2', 'Collections by Method',
                        list(by_method.keys()), list(by_method.values()), '_chart_method')
    if by_day:
        days_sorted = sorted(by_day.keys())
        _add_line_chart(wb, ws, 'H20', 'Daily Collections', days_sorted,
                         [by_day[d] for d in days_sorted], '_chart_daily', y_title='Ksh')
    if by_officer:
        _add_bar_chart(wb, ws, 'H38', 'Collections by Loan Officer', list(by_officer.keys()),
                        [('Amount (Ksh)', list(by_officer.values()))], '_chart_officer', y_title='Ksh')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='collections_report.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/api/export/collections/csv')
@login_required
def export_collections_csv():
    """CSV sibling of /api/collection-report -- same date_from/date_to filters
    and query, so the download always matches what's on screen."""
    date_from = request.args.get('date_from', date.today().replace(day=1).isoformat())
    date_to = request.args.get('date_to', date.today().isoformat())

    repayments = get_db().execute(
        """SELECT repayments.*, loans.loan_number FROM repayments
           LEFT JOIN loans ON loans.id = repayments.loan_id
           WHERE repayments.payment_date >= %s AND repayments.payment_date <= %s
           ORDER BY repayments.payment_date DESC""",
        (date_from, date_to)
    ).fetchall()

    headers = ['Loan No', 'Amount', 'Payment Method', 'Payment Date', 'Reference']

    rows = [
        [
            r['loan_number'], r['amount'], r['payment_method'],
            r['payment_date'], r['reference_number'] or '',
        ]
        for r in repayments
    ]

    return _csv_response(headers, rows, 'collections_report.csv')


@reports_bp.route('/api/export/regional/csv')
@login_required
def export_regional_csv():
    data = _compute_regional_performance(request.args.get('month'))
    return _csv_response(_REGIONAL_HEADERS, _regional_rows(data['regions']),
                          f"regional_performance_{data['month']}.csv")
